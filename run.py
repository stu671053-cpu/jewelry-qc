#!/usr/bin/env python3
"""
域骉控股珠宝质检数据分析 — 主入口

用法:
  python run.py                          # 用默认路径运行
  python run.py /path/to/input.xlsx      # 指定输入文件
  python run.py --test                   # 只运行Golden测试
  python run.py --test /path/to/input.xlsx  # 运行测试 + 处理数据

流程: Golden测试通过 → 读取Excel → 规则引擎 → 输出标注
"""

import sys
import os
from pathlib import Path
from datetime import datetime

# 将项目根目录加入 sys.path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from engine import QCEngine, verify_golden_tests, verify_isolation_tests


DEFAULT_INPUT = "/Users/yubiao/Desktop/珠宝对照分析-R6完整版_已标注.xlsx"


def main():
    # 解析参数
    args = sys.argv[1:]
    test_only = "--test" in args
    input_path = DEFAULT_INPUT

    for arg in args:
        if arg != "--test" and not arg.startswith("-"):
            input_path = arg

    # 1. 初始化引擎
    print("=" * 60)
    print("域骉控股珠宝质检数据分析")
    print("=" * 60)

    engine = QCEngine()
    print(f"已加载规则: {[r['name'] for r in engine.rules]}")

    # 2. 运行Golden测试
    print("\n--- Golden 测试 ---")
    golden_ok = verify_golden_tests(engine)

    if not golden_ok:
        print("\nGolden 测试未通过! 请修复后再运行生产数据。")
        sys.exit(1)

    # 3. 运行隔离性测试
    print("\n--- 规则隔离性测试 ---")
    isolation_ok = verify_isolation_tests(engine)

    if not isolation_ok:
        print("\n隔离性测试未通过! 存在规则间字段耦合风险，请检查。")
        sys.exit(1)

    if test_only:
        print("\n所有测试全部通过，退出。")
        sys.exit(0)

    # 4. 读取Excel (使用openpyxl，避免numpy依赖)
    print(f"\n--- 读取数据 ---")
    print(f"文件: {input_path}")
    if not os.path.exists(input_path):
        print(f"文件不存在: {input_path}")
        sys.exit(1)

    from openpyxl import load_workbook

    wb = load_workbook(input_path, read_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 读取数据到列表
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))
    wb.close()

    print(f"数据行数: {len(data)}, 列数: {len(headers)}")

    # 5. 运行规则引擎
    print("\n--- 运行规则引擎 ---")
    start_time = datetime.now()

    # 为每行数据应用规则
    for row in data:
        results = engine.apply_row(row)
        # 将结果添加到行数据
        row.update(results)

    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"处理完成，耗时: {elapsed:.2f}秒")

    # 6. 输出统计
    print("\n--- 检查结果统计 ---")
    for rule in engine.rules:
        col = rule["column"]
        if col in data[0]:
            # 统计结果
            from collections import Counter
            vc = Counter(row[col] for row in data)
            total = len(data)
            normal = vc.get("正常", 0) + vc.get("正确", 0) + vc.get("质检通过", 0) + vc.get("备注无误", 0)
            empty = vc.get("", 0) + vc.get("未检查", 0)
            abnormal = total - normal - empty
            print(f"  {col}: 异常{abnormal}条 / 正常{normal}条 / 空{empty}条")

    # 7. 保存到 WorkBuddy 工作区（避免沙箱权限问题）
    input_stem = Path(input_path).stem
    output_path = Path("/Users/yubiao/WorkBuddy/2026-06-05-10-29-15") / f"{input_stem}_已标注.xlsx"

    # 使用openpyxl写入结果
    from openpyxl import Workbook

    wb_out = Workbook()
    ws_out = wb_out.active

    # 写入表头
    all_headers = headers + [rule["column"] for rule in engine.rules]
    ws_out.append(all_headers)

    # 写入数据
    for row in data:
        ws_out.append([row.get(h, "") for h in all_headers])

    wb_out.save(output_path)
    print(f"\n结果已保存至: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
